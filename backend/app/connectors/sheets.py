import io
import logging
import time
from datetime import datetime

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from app.connectors.base import BaseConnector, ConnectorStatus
from app.processing.pipeline import RawDocument

logger = logging.getLogger(__name__)

SCOPES_SHEETS = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]
MAX_ROWS = 10_000


def _execute_with_retry(request, max_retries=5, initial_delay=2.0):
    """Execute Google API request with exponential backoff on HTTP 429 / 5xx."""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            return request.execute(num_retries=3)
        except HttpError as e:
            if e.resp.status in (429, 500, 502, 503, 504) and attempt < max_retries - 1:
                logger.warning(
                    f"Google API rate limit / error (HTTP {e.resp.status}). Retrying in {delay:.1f}s... (attempt {attempt + 1}/{max_retries})"
                )
                time.sleep(delay)
                delay *= 2
            else:
                raise


class SheetsConnector(BaseConnector):
    async def authenticate(self, **kwargs) -> dict:
        return {}

    def _build_services(self, token_dict: dict):
        creds = Credentials(
            token=token_dict.get("access_token"),
            refresh_token=token_dict.get("refresh_token"),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=token_dict.get("client_id"),
            client_secret=token_dict.get("client_secret"),
            scopes=SCOPES_SHEETS,
        )
        sheets_svc = build("sheets", "v4", credentials=creds)
        drive_svc = build("drive", "v3", credentials=creds)
        return sheets_svc, drive_svc

    async def fetch_documents(self, connector_record, since: datetime | None = None) -> list[RawDocument]:
        import asyncio
        from app.connectors.utils import decrypt_token
        token_dict = decrypt_token(connector_record.oauth_token_encrypted)
        return await asyncio.to_thread(self._fetch_documents_sync, token_dict, since)

    def _fetch_documents_sync(self, token_dict: dict, since: datetime | None) -> list[RawDocument]:
        sheets_svc, drive_svc = self._build_services(token_dict)

        query = "mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"
        if since:
            since_str = since.strftime("%Y-%m-%dT%H:%M:%S")
            query += f" and modifiedTime > '{since_str}'"

        request = drive_svc.files().list(
            q=query,
            fields="files(id, name, modifiedTime, owners, webViewLink)",
            pageSize=100,
        )
        response = _execute_with_retry(request)

        results = []
        files = response.get("files", [])
        for index, sheet_file in enumerate(files):
            # Pace requests (1 second delay per spreadsheet) to comfortably stay within 60 req/min limit
            if index > 0:
                time.sleep(1.0)

            try:
                docs = self._process_sheet(sheets_svc, sheet_file)
                results.extend(docs)
            except Exception as e:
                logger.error(f"Failed to process Google Sheet '{sheet_file.get('name', sheet_file.get('id'))}': {e}")
                time.sleep(2.0)

        return results

    def _process_sheet(self, sheets_svc, sheet_file: dict) -> list[RawDocument]:
        req = sheets_svc.spreadsheets().get(spreadsheetId=sheet_file["id"])
        spreadsheet = _execute_with_retry(req)

        results = []
        for sheet in spreadsheet.get("sheets", []):
            props = sheet.get("properties", {})
            tab_name = props.get("title", "Sheet")

            # Small delay per tab call to avoid spiking read rate limits
            time.sleep(0.5)

            val_req = sheets_svc.spreadsheets().values().get(
                spreadsheetId=sheet_file["id"],
                range=tab_name,
            )
            try:
                values_response = _execute_with_retry(val_req)
            except Exception as e:
                logger.error(f"Failed to fetch values for tab '{tab_name}' in sheet '{sheet_file.get('name')}': {e}")
                continue

            rows = values_response.get("values", [])
            if not rows:
                continue

            rows = rows[:MAX_ROWS]
            if len(rows) == MAX_ROWS:
                rows.append(["[Row limit reached — first 10,000 rows indexed]"])

            header = rows[0] if rows else []
            text_lines = ["\t".join(str(c) for c in header)]

            for idx, row in enumerate(rows[1:], start=2):
                row_parts = []
                for col_idx, val in enumerate(row):
                    col_name = header[col_idx] if col_idx < len(header) else f"Col{col_idx+1}"
                    row_parts.append(f"{col_name}={val}")
                text_lines.append(f"Row {idx}: " + ", ".join(row_parts))

            text = "\n".join(text_lines)
            modified = sheet_file.get("modifiedTime")
            dt = datetime.fromisoformat(modified.replace("Z", "+00:00")) if modified else None

            results.append(
                RawDocument(
                    source_id=f"{sheet_file['id']}_{tab_name}",
                    title=f"{sheet_file['name']} — {tab_name}",
                    text=text,
                    author=sheet_file.get("owners", [{}])[0].get("emailAddress", ""),
                    source_url=sheet_file.get("webViewLink"),
                    source_created_at=dt,
                    source_type="spreadsheet",
                )
            )

        return results

    async def get_status(self, connector_record) -> ConnectorStatus:
        return ConnectorStatus(
            status=connector_record.status,
            last_sync_at=connector_record.last_sync_at,
            document_count=connector_record.document_count,
            error_message=connector_record.error_message,
        )


def process_excel_file(filename: str, data: bytes) -> list[RawDocument]:
    results = []
    try:
        if filename.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True, max_row=MAX_ROWS))
                text = _rows_to_text(rows)
                results.append(
                    RawDocument(
                        source_id=f"excel_{filename}_{sheet_name}",
                        title=f"{filename} — {sheet_name}",
                        text=text,
                        source_type="spreadsheet",
                    )
                )
        elif filename.endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(file_contents=data)
            for sheet in wb.sheets():
                rows = [tuple(sheet.row_values(i)) for i in range(min(sheet.nrows, MAX_ROWS))]
                text = _rows_to_text(rows)
                results.append(
                    RawDocument(
                        source_id=f"excel_{filename}_{sheet.name}",
                        title=f"{filename} — {sheet.name}",
                        text=text,
                        source_type="spreadsheet",
                    )
                )
    except Exception:
        pass

    return results


def _rows_to_text(rows: list) -> str:
    if not rows:
        return ""
    header = [str(c) for c in rows[0]]
    lines = ["\t".join(header)]
    for idx, row in enumerate(rows[1:], start=2):
        parts = []
        for col_idx, val in enumerate(row):
            col_name = header[col_idx] if col_idx < len(header) else f"Col{col_idx+1}"
            parts.append(f"{col_name}={val}")
        lines.append(f"Row {idx}: " + ", ".join(parts))
    return "\n".join(lines)
